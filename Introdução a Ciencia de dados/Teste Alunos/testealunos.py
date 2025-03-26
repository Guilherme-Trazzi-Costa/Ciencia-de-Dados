from openpyxl import load_workbook

arquivo = load_workbook("Alunos.xlsx")

#print(arquivo.sheetnames)

aba_atual = arquivo.active


aba_alunos = arquivo["Planilha1"]
print(aba_alunos)

#Selecionar e alterar valor de celula
valor_a1 = aba_alunos["A1"].value
valor_b1 = aba_alunos.cell(row=1, column = 2).value

aba_alunos.cell(row=1, column = 2).value = "Prova 1"

arquivo.save("Alunos2.xlsx")

#Ultima linha do excel
print(aba_alunos.max_row)


